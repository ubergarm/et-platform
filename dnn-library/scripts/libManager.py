#!/usr/bin/env python3
import os
import re
import argparse
import sys
from string import Template
import functools

from openpyxl import load_workbook, Workbook
from openpyxl.styles import colors, PatternFill, Color, Font
from openpyxl.utils import get_column_letter

class OperatorsEnum:
    def __init__(self, hostswdir):
        fname = os.path.join ( hostswdir, 'Neuralizer/inc/IOperators.inc');
        regexp = re.compile(r'\s*IOPERATOR_ENUM\((.*)\)\s*')
        self.__enum = []
        with open(fname) as f:
            for line in f:
                m = regexp.match( line )
                if m:
                    self.__enum.append("ET_" + m.group(1))

    def get(self):
        return self.__enum
        
class LibManagerSheet:
    ################################################################################
    # constants
    ################################################################################
    headerKeys = ("Operator", "nrOutTensors", "nrInTensors", "members", "templateElk","extraImpl", "implSel")
    instHeader=("Operator", "nrOutTensors", "nrInTensors", "templateElk")

    headerFill = PatternFill(fgColor = "DDDDDD", fill_type = "solid")
    #    operatorFill = [PatternFill(fgColor = "3FBE7E", fill_type = "solid"),
    #                    PatternFill(fgColor = "5F9E9F", fill_type = "solid") ]
    operatorFill = [None]

    memberTypeMap = { "HalfWindowSize": "uint32_t",
                      "Alpha": "float",
                      "Beta": "float",
                      "K": "float",
                      "TopK": "uint32_t",
                      "Divisor": "uint64_t",
                      "SignFollowDivisor": "bool",
                      "Axis": "dim_t",
                      "KeepDims": "bool",
                      "Kernels": "std::array<uint32_t, default_kernels_size> &",
                      "Strides": "std::array<uint32_t, default_kernels_size> &",
                      "Pads": "std::array<uint32_t, default_kernels_size> &",
                      "Group": "uint32_t",
                      "Offsets": "dim_array_t &",
                      "Shuffle": "std::array<uint32_t, max_tensor_dimensions> &",
                      "Mask": "std::array<uint64_t, default_mask_size>&",
                      "BatchDims": "uint32_t",
                      "Count": "uint32_t",
                      "SyncOffset": "uint32_t",
                      "Value": "float",
                      "Exclusive": "bool",
                      "Reverse": "bool",
                      "BlockSize": "uint32_t",
                      "Axes": "std::array<uint32_t, default_axes_size> &",
                      "RszScale": "std::array<float, default_rszscale_size> &",
                      "HasEndOffset": "bool",
                      "Transposed": "bool",
                      "TensorsAligned": "bool",
                      "RszScale": "std::array<float, default_rszscale_size> &",
                      "Dilation": "uint32_t",
                      "CenterPointBox": "int64_t",
                      "MaxOutputBoxesPerClass": "int64_t",
                      "IouThreshold": "float",
                      "ScoreThreshold": "float",
                      "IsTFVersion4": "bool"
                    }

    # members that end up adding another template paramer (they are std::array<T, N>)
    # assuming all the ones with the same value (e.g. Kernels and Strides share the same template param
    # if not, just rename one of them
    memberExtraTpl = { "Kernels": "size_t KN",
                       "Strides":  "size_t KN",
                       "Pads": "size_t PN",
                       "Shuffle": "size_t N",
                       "Mask": "size_t MN",
                       "Axes": "size_t N",
                       "RszScale": "size_t RSZN"
    }


    ################################################################################
    # constructor
    ################################################################################

    def __init__(self, spreadsheet, hostswdir = None, glowdir = None):

        enum = OperatorsEnum(hostswdir)
        self.__enum = enum.get()
        
        try:
            self.__wb = load_workbook(spreadsheet, data_only=True)
            self._existing = True
        except FileNotFoundError:
            self._existing = False

        if self._existing:
            self.__codeGen(hostswdir)
        else:
            self.__xlsGen(spreadsheet, hostswdir, glowdir)

    ################################################################################
    # methods to create the spread sheet skeleton
    ################################################################################

    def __xlsGen(self, spreadsheet, hostswdir, glowdir):
        # parse instruction generator
        insts = InstructionGenParser(glowdir)

        #find implementations in dnn_lib dir
        implementations = self.__findImplementations(hostswdir + "/dnn_lib/include/inlining")

        
        print ("Excel not found: creating. Edit if necessary and rerun.")
        self.__row = 1
        self.__wb = Workbook()
        self.__ws = self.__wb.active
        self.__ws.title = "LibManager"
        
        # header
        self.__addRow(LibManagerSheet.headerKeys, LibManagerSheet.headerFill)

        # add operators
        for op in sorted(insts.get_all()):
            conf = insts.get(op)
            
            # count tensor operands            
            nrIn = 0
            nrOut = 0
            for i in conf['operands']:
                if i['kind'] == 'OperandKind::Out' or i['kind'] == 'OperandKind::InOut':
                    nrOut+=1
                elif i['kind'] == 'OperandKind::In':
                    nrIn+=1
                else:
                    print("in %s, tensor operand kind %s not supported: %s will be ignored" % (op, i['kind'], i["name"]), file = sys.stderr)

            # list non tensor operands
            members = [ i['name'] for i in conf['members']]

            template_params = "" #empty, will have to be added manually in the spreasheet
            
            if op not in implementations:
                print("no implementations found for %s: skipping" % op, file = sys.stderr)
                continue
                
            extraImpl = [ k for k in implementations[op] ]
                
            values = [ op, nrOut, nrIn,
                       ', '.join(members),
                       template_params,
                       ', '.join(extraImpl),
                       "default"]
            
            self.__addRow(values, LibManagerSheet.operatorFill[self.__row % len(LibManagerSheet.operatorFill) ])

            
        # and adjust width to fit contents
        self.__adjustColumnWidth()

        # and finally save
        self.__wb.save(filename = spreadsheet)

    def __addRow(self, values, fill):
        for i, val in enumerate (values):
            cell = self.__ws.cell(self.__row, 1 + i, val)
            if fill != None:
                cell.fill = fill
        self.__row+=1

    def __adjustColumnWidth(self):
        for column_cells in self.__ws.columns:
            col = column_cells[0].column
            col_letter = get_column_letter(col)
            fit = 6 +  max(len(str(cell.value)) for cell in column_cells)                
            self.__ws.column_dimensions[col_letter].width = fit


    def __findImplementations(self, baseDir):
        implementations={}
        reHeader = re.compile(r'.*\.h')
        reImpl = re.compile(r'.*void.*fwdLib(.*)Inst(.*)\(.*')
        eltWiseImpl = re.compile(r'\s*EltWiseInst\((.*), .*')
        
        files = [os.path.join(baseDir, f) for f in os.listdir(baseDir) if os.path.isfile(os.path.join(baseDir, f)) and reHeader.match(f)]        
        for fname in files:
            found = 0
            with open(fname) as f:
                for line in f:
                    m = reImpl.match(line)
                    if m:
                        op = m.group(1)
                        details = m.group(2)
                        if op not in implementations:
                            implementations[op] = []
                        if len(details) > 0:
                            implementations[op].append(details)
                        continue
                    # special way elementwise instructions are defined
                    m = eltWiseImpl.match(line)
                    if m:
                        op = m.group(1)
                        implementations[op] = ['Threaded', 'Vectorized']
                        
        return implementations
    
    ################################################################################
    #  Methods to retrieve information from the sheet
    ################################################################################
        
    def __codeGen(self, hostswdir):
        # load configuration from spreadsheet
        self.loadSheet()
        # create LibApi table
        self.genLibApi(hostswdir)
        # create non-inline functions and extern template definitions
        self.genNonInline(hostswdir)

    def parseLibManagerSheet(self, row):
        operator = { "gen": False}
        for h,cell in zip(LibManagerSheet.headerKeys, row):
            operator[h] = cell.value
            
        if operator["Operator"][0] == "#": #disabled from sheet
            return
        
        enum = "ET_" + operator["Operator"].lower()
        self.__configs[enum] = operator
        
    def parseInstancesSheet(self, row):
        instances= []
        for i,cell in enumerate(row):
            if i == 0: # op name
                op = cell.value
            elif i >= len(LibManagerSheet.instHeader):
                v = cell.value
                if v:
                    if v[0] != "#": #ignore commented out specializations
                        instances.append(v)
        if op == 0 or op[0] == "#" : #ignoring empty or disabled
            return

        enum = "ET_" + op.lower()                    
        self.__configs[enum]["instances"] = instances

    def loadSheet(self):

        # read the sheets
        self.__configs = {}
        parser = [self.parseLibManagerSheet, self.parseInstancesSheet ]
        for i,sn in enumerate(("LibManager","Instances")):
            ws = self.__wb[sn]
            skipHeader  = True
            for row in ws.rows:
                if skipHeader:
                    skipHeader = False
                    continue
                parser[i](row)
        

    def genLibApi(self, hostswdir):
        # generate table for header file
        table = [ self.tableEntry(i) for i in self.__enum ] 
        tableStr = self.formatTable(table)
       
        # check for entries in the header file that have not been used
        missing = [i for i in self.__configs if not self.__configs[i]["gen"]]
        for i in missing:
            print("Spreadhseet row for %s not used" % i, file = sys.stderr)

        # and create output
        self.outputLibApi(hostswdir, tableStr)
            
    def tableEntry(self, op):
        if op in self.__configs:
            conf = self.__configs[op];
            conf["gen"] = True
            members = []
            versions = []
            template = 0
            
            if conf["members"]:
                members = ["mb" + i.replace(" ", "") for i in conf["members"].split(',')]
                
            if conf["extraImpl"]:
                versions = ['"' + i.replace(" ", "") + '"' for i in conf["extraImpl"].split(',')]

            #filter out commented implementations
            versions = [i for i in versions if i[1] != '#']
            
            if conf["templateElk"] == None:
                raise Exception("empty tenplate definition for %s. Use NONE if the fnc doesn't use templates" % op)
            elif conf["templateElk"] != "NONE":
                template = functools.reduce( lambda a,b : a | (1 << int(b)), str(conf["templateElk"]).split(','), 0) ;
                
            if conf["implSel"] == "default":
                implSel = "false"
            elif conf["implSel"] == "custom":
                implSel = "true"
            else:
                raise Exception("implSel has to be either 'default' or 'custom' for op %s" % op)
                
            return { "enum": op,
                     "name" : conf["Operator"],
                     "nrOutputTensors": conf["nrOutTensors"],
                     "nrInputTensors": conf["nrInTensors"],
                     "members": "{%s}" % ", ".join(members),
                     "template": template,
                     "versions":  "{%s}" % ", ".join(versions),
                     "implSel": implSel
            }

        else:
            print("WARN: Could not find spreadsheet row for %s" % op, file = sys.stderr)
            return {"enum": op,
                    "name" : "notImplemented",
                    "nrOutputTensors": 0,
                    "nrInputTensors": 0,
                    "members": "{}",
                    "template": 0,
                    "versions": "{}",
                    "implSel": "false"
            }


    def formatTable(self, table):
        s = Template('''     /**** $enum ****/
     { "$name", // name
       $nrOutputTensors, // # outs
       $nrInputTensors,  // # ins
       $members, // members
       $template, // template param mask
       $versions, // impl versions
       $implSel // custom impl selector
     }''')
        entries = [ s.substitute(e) for e in table]        
        return ",\n".join(entries)

    def outputLibApi(self, hostswdir, tableStr):
        contents = []
        startMark = re.compile(r'\s*// INSTR_CONFIG_TABLE_BEGIN')
        endMark = re.compile(r'\s*// INSTR_CONFIG_TABLE_END')
        fname = os.path.join ( hostswdir, 'dnn_lib/include/LibApi.h');
        found = False
        written = False
        with open(fname) as f:
            for line in f:
                if not found:
                    contents.append(line)
                    if startMark.match(line):
                        contents.append(tableStr + "\n")
                        written = True
                        found = True
                else:
                    if endMark.match(line):
                        contents.append(line)
                        found = False

        if found:
            raise Exception("didn't not find end of table in %s" % fname)
        if not written:
            raise Exception("didn't not find start of table in %s" % fname)

        with open(fname, "w") as f:
            f.writelines(contents)

    def genNonInline(self, hostswdir):
        fncs = {}
        for op in self.__enum:
            if op in self.__configs:
                opData = []
                conf = self.__configs[op]
                if not conf["gen"]:
                    continue

                inst = conf["instances"]
                versions = [""]
                if conf["extraImpl"]:
                    versions += [i.replace(" ", "")  for i in conf["extraImpl"].split(',')]
                #filter out commented implementations
                versions = [i for i in versions if len(i) == 0 or i[0] != '#']
                
                members = []
                if conf["members"]:
                    members = [ i.replace(" ", "") for i in conf["members"].split(',')]

                for v in versions:
                    fname = "fwdLib%sInst%s" % (conf["Operator"], v)
                    if len(inst) == 0:
                        opData.append(self.getImplInfo(conf["Operator"], fname, conf, members))
                    else:
                        for tpl in inst:
                            opData.append(self.getImplInfo(conf["Operator"],fname, conf, members, tpl))
                fncs[op] = opData

        # and generate the code
        self.genLibNodes(hostswdir, fncs)
        self.genCppNodes(hostswdir, fncs)

    def genLibNodes(self, hostswdir, fncs):
        #create libNodes.h
        autogenMsg = "// File automatically generated with:\n//  %s\n//  cwd=%s\n" % (' '.join(sys.argv), os.getcwd())
        
        hFile = os.path.join ( hostswdir, 'dnn_lib/include/LibNodes.h');
        code = []
        for op in fncs:
            code+= [ "\n/****************************************************************************",
                     "*  %s implementations" % fncs[op][0]['opname'],
                     "****************************************************************************/",
                     "// declarations"]
            declared = {} # keep track of what has been declared (not to repeat several times the same
                          # declarations) => required because splat functions have different params
                          # depending on the templates
            for i in fncs[op]:
                decl = "%s\nvoid %s(%s);" % (i['templateDecl'], i['fname'], i['callDeclHeader'])
                if decl not in declared:
                    code.append(decl)
                    declared[decl] = True

            code.append("\n// extern template declarations")
            for i in fncs[op]:
                if len(i['templateInst']) > 0:
                    code.append("extern template void %s%s(%s);" % (i['fname'], i['templateInst'], i['callDecl']))

        code = "\n".join(code)
        with open(hFile, "w") as f:
            f.write("""%s
#ifndef LIBNODES_H_
#define LIBNODES_H_

#include "LibTensor.h"
#include "inlining.h"

namespace dnn_lib {
static constexpr size_t default_kernels_size = 2;
static constexpr size_t default_mask_size = max_tensor_dimensions;
static constexpr size_t default_axes_size = max_tensor_dimensions;
static constexpr size_t default_rszscale_size = max_tensor_dimensions;
%s
} // namespace dnn_lib

#endif /* LIBNODES_H_ */
""" % (autogenMsg, code))

    def genCppNodes(self, hostswdir, fncs):
        for op in fncs:
            opname = fncs[op][0]['opname']
            cppFile = os.path.join(hostswdir, "dnn_lib/src/%sInst.cpp" % opname )
            with open(cppFile, "w") as f:
                f.write("""
#include "LibNodes.h"
 
namespace dnn_lib {
  ////////////////////////////////////////////////////////////////////////////////
  // Forward call to corresponding dnn_lib::inlining implementations
  ////////////////////////////////////////////////////////////////////////////////
 """)
                created = {}
                for i in fncs[op]:
                    fnc = "%s\n  void %s(%s)" % (i['templateDecl'], i['fname'], i['callDecl'])
                    if fnc not in created:
                        f.write("""
  %s
  {
    dnn_lib::inlining::%s%s(%s);
  }
"""
                                % (fnc, i['fname'], i['templateFwd'], i['callInst']))
                        created[fnc] = True
                        f.write("""
  ////////////////////////////////////////////////////////////////////////////////
  // Template specializations (declared with 'extern template' in LibNodes.h)
  ////////////////////////////////////////////////////////////////////////////////
""")
                created = {}
                for i in fncs[op]:
                    decl = "template void %s%s(%s);\n" % (i['fname'], i['templateInst'], i['callDecl'])
                    if fnc not in created and len(i['templateInst']) > 0:
                        f.write(decl);
                        created[decl] = True

                f.write("} // dnn_lib\n")


        
        
    def getImplInfo(self, opname, fname, conf, members, tpl = None):
        if conf["templateElk"] == "NONE":
            tensorTpl = []
        else:
            tensorTpl = [ int(i) for i in str(conf["templateElk"]).split(',')]
        tplInst = "" if tpl == None else "<%s>" % tpl
        info = { 'fname': fname,
                 'opname': opname,
                 'templateDecl' : [],
                 'templateInst': tplInst,
                 'templateFwd': [],
                 'callDecl': [],
                 'callInst':[]}

        
        for i in range(conf["nrOutTensors"]):
            info['callDecl'].append("LibTensor* out%d" % i)
            info['callInst'].append("out%d" % i)            
            if i in tensorTpl:
                info['templateDecl'].append('ElemKind out%dType' % i)
                info['templateFwd'].append('out%dType' % i)                
                
        for i in range(conf["nrInTensors"]):
            info['callDecl'].append("LibTensor* in%d" % i)
            info['callInst'].append("in%d" % i)            
            if i + conf["nrOutTensors"] in tensorTpl:
                info['templateDecl'].append('ElemKind in%dType' % i)
                info['templateFwd'].append('in%dType' % i)                

            
        for i in members:
            info['callDecl'].append("const %s %s" % (self.memberType(i, tpl), i))
            info['callInst'].append(i)
            ##if i in LibManagerSheet.memberExtraTpl:
            ##    t = LibManagerSheet.memberExtraTpl[i]
            ##    info['templateDecl'].append( "%s %s" %(t,i))

        info['callDeclHeader'] = info['callDecl'].copy() # same as callDecl, but with default values
        info['callDecl']+=[ "const uint64_t flags",
                            "const uint32_t minionOffset",
                            "const uint32_t assignedMinions" ]
        info['callDeclHeader']+=[ "const uint64_t flags",
                                  "const uint32_t minionOffset = 0",
                                  "const uint32_t assignedMinions = 0" ]        
        info['callInst']+=["flags", "minionOffset", "assignedMinions"]
                               
        # convert to strings
        info['callDecl'] = ', '.join(info['callDecl'])
        info['callDeclHeader'] = ', '.join(info['callDeclHeader'])
        info['callInst'] = ', '.join(info['callInst'])
        if len(info['templateDecl']) == 0:
            info['templateDecl'] = ""
            info['templateFwd'] = ""
        else:
            info['templateDecl'] = "template <%s>" % (', '.join(info['templateDecl']))
            info['templateFwd'] = "<%s>" % (', '.join(info['templateFwd']))

        return info

    def memberType(self, m, tpl = None):
        # special case: "Value" => type is float except int64_t if first param is int64_t
        if m == "Value" and re.match(r'int64_t', tpl):
            return "int64_t"
        
        # expected exception if member not found
        return LibManagerSheet.memberTypeMap[m]


        
if __name__ == "__main__":
    # parse command line options
    parser = argparse.ArgumentParser("Create Operator test")
    parser.add_argument("--swplatform-root", help="sw-platform root dir", required = True) 
    parser.add_argument("--excel", help="Excel file to use", required = True)
    args = parser.parse_args()
    sys.path.append(os.path.join (args.swplatform_root, 'scripts/testing/operatorTests/'))
    from instructionGenParser import InstructionGenParser

    hostsw_dir = os.path.join (args.swplatform_root,'host-software/host-sw/')
    glow_dir = os.path.join (args.swplatform_root,'host-software/glow/')
    sheet = LibManagerSheet(args.excel, hostsw_dir, glow_dir)

